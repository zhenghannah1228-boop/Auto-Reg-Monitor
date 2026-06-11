"""
matcher.py — 推文解读与法规库的双向匹配核心库
被 compile_tweet.py 和 build_kanban.py 共用。
"""
import openpyxl, json, re, glob, os

PROJECT_DIR = os.environ.get("REG_LIB_DIR") or (
    "/mnt/project" if os.path.isdir("/mnt/project")
    else os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "output", "library_linked"))

# ---- 规范化法规编号，用于跨文件匹配 ----
def norm(s):
    if not s:
        return ''
    s = str(s).upper()
    s = (s.replace('UN REGULATION NO.', 'UN R')
           .replace('UN ECE', 'UN R')
           .replace('REGULATION NO.', 'R')
           .replace('REGULATION (EU)', 'EU')
           .replace('(EU)', 'EU'))
    s = re.sub(r'\s+', '', s)
    return s

# ---- 剥离系列号/修订号后缀，得到基础编号 ----
# UN R100.04 / UN R79-04 / UN R79-04REVISION5 / UN R13H附件15 → UN R100 / UN R79 / UN R13
# 保留 R13H 这类字母变体（H 是法规本体的一部分），只剥数字系列后缀
def base_key(match_key):
    nk = norm(match_key)
    # UN R<数字>[字母变体H/HZ] 后面的 .NN / -NN / REV / AMEND / 附件 等一律剥掉
    m = re.match(r'^(UNR\d+[A-Z]?)', nk)
    if m:
        return m.group(1)
    # GB / GBT 后跟数字，剥掉 -YYYY 年份
    m = re.match(r'^(GB[T]?\d+)', nk)
    if m:
        return m.group(1)
    # EU YYYY/NNNN 保持完整（年份是编号本体）
    m = re.match(r'^(EU\d{4}/\d+)', nk)
    if m:
        return m.group(1)
    # FMVSS / ADR / CONTRAN 等
    m = re.match(r'^([A-Z]+\d+)', nk)
    if m:
        return m.group(1)
    return nk

# ---- 扫描法规库，建立 {规范化编号: [(文件,sheet,行,原编号,中文名)]} 索引 ----
def build_library_index(lib_files=None):
    if lib_files is None:
        lib_files = glob.glob(os.path.join(PROJECT_DIR, "*法规清单.xlsx"))
    index = {}
    for f in lib_files:
        label = os.path.basename(f).replace('法规清单.xlsx', '').replace('汽车', '')
        try:
            wb = openpyxl.load_workbook(f, read_only=True)
        except Exception:
            continue
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
            header = [str(c) if c else '' for c in next(rows, [])]
            if '法规编号' not in header:
                continue
            ci = header.index('法规编号')
            ni = header.index('法规中文名称') if '法规中文名称' in header else None
            for ridx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if ci < len(row) and row[ci]:
                    regno = str(row[ci])
                    cname = str(row[ni]) if ni is not None and ni < len(row) and row[ni] else ''
                    index.setdefault(norm(regno), []).append(
                        {"file": label, "sheet": sheet, "row": ridx,
                         "reg_no": regno, "cn_name": cname})
        wb.close()
    return index

# ---- 用 match_key 在索引里查匹配（带数字边界，避免 R12 命中 R121）----
# 三级回退：精确编号 → 边界前缀 → 基础编号（剥系列/修订后缀）
def match_key_to_library(match_key, index):
    nk = norm(match_key)
    if not nk:
        return []
    # 1) 精确命中
    if nk in index:
        return index[nk]
    # 2) 边界前缀匹配
    hits = []
    for k, v in index.items():
        if k == nk:
            hits.extend(v); continue
        if k.startswith(nk):
            tail = k[len(nk):]
            if tail == '' or not tail[0].isdigit():
                hits.extend(v)
    if hits:
        return hits
    # 3) 回退到基础编号（处理 UN R100.04 / UN R79-04 这类带系列号的）
    bk = base_key(match_key)
    if bk != nk:
        if bk in index:
            return index[bk]
        for k, v in index.items():
            if k.startswith(bk):
                tail = k[len(bk):]
                if tail == '' or not tail[0].isdigit():
                    hits.extend(v)
    return hits

# ---- 给一批 insights 计算 library 链接，写回 _matched 字段 ----
def annotate_matches(insights, index):
    for it in insights:
        for reg in it.get('regulations', []):
            hits = match_key_to_library(reg['match_key'], index)
            reg['_matched'] = hits
            reg['_matched_count'] = len(hits)
    return insights

if __name__ == '__main__':
    idx = build_library_index()
    print(f"法规库索引完成：{len(idx)} 个唯一编号")
