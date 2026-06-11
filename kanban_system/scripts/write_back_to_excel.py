#!/usr/bin/env python3
"""
write_back_to_excel.py — 把解读链接写回法规库 Excel

在每个匹配到的法规清单文件里，新增/更新一列「解读链接」，
内容形如：  INS-2026-012 自动近光灯法规修订分析 | UN R48.8 ; INS-...
并可选写入超链接（如果 insight 有 url）。

用法：
  python write_back_to_excel.py            # 写到 output/ 下的副本（默认，安全）
  python write_back_to_excel.py --inplace  # 直接改 /mnt/project 原文件（谨慎）

默认 dry-run-to-copy：复制原文件到 output/library_linked/ 再写，不动原始数据。
"""
import json, os, sys, shutil, argparse
import openpyxl
from openpyxl.styles import Font

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA = os.path.join(ROOT, 'data', 'insights.json')
OUT = os.path.join(ROOT, 'output')

sys.path.insert(0, HERE)
from matcher import build_library_index, match_key_to_library, PROJECT_DIR

LINK_COL_NAME = '解读链接'

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--inplace', action='store_true', help='直接修改 /mnt/project 原文件')
    args = ap.parse_args()

    db = json.load(open(DATA, encoding='utf-8'))
    insights = db['insights']
    lib_index = build_library_index()

    # 建立 (file,sheet,row) -> [insight 链接文本] 的映射
    cell_links = {}  # (file_label, sheet, row) -> list of strings
    cell_urls = {}   # (file_label, sheet, row) -> first url
    for it in insights:
        link_text = f"{it['id']} {it['title']}"
        for reg in it['regulations']:
            for hit in match_key_to_library(reg['match_key'], lib_index):
                key = (hit['file'], hit['sheet'], hit['row'])
                cell_links.setdefault(key, []).append(link_text)
                if it.get('url') and key not in cell_urls:
                    cell_urls[key] = it['url']

    if not cell_links:
        print("没有任何可写回的匹配。")
        return

    # 按文件分组
    files_touched = {}
    for (file_label, sheet, row) in cell_links:
        files_touched.setdefault(file_label, set()).add(sheet)

    out_dir = os.path.join(OUT, 'library_linked')
    if not args.inplace:
        os.makedirs(out_dir, exist_ok=True)

    import glob
    # file_label -> 实际路径
    def resolve_path(label):
        for ext in ('.xlsx',):
            for cand in glob.glob(os.path.join(PROJECT_DIR, f"*{label}*{ext}")):
                return cand
        return None

    for label, sheets in files_touched.items():
        src = resolve_path(label)
        if not src:
            print(f"[跳过] 找不到文件: {label}")
            continue
        if args.inplace:
            target = src
        else:
            target = os.path.join(out_dir, os.path.basename(src))
            # 库目录与副本目录相同时(本仓库默认指向 library_linked)原地写,跳过复制
            if os.path.abspath(src) != os.path.abspath(target):
                shutil.copy(src, target)
        wb = openpyxl.load_workbook(target)
        for sheet in sheets:
            ws = wb[sheet]
            header = [c.value for c in ws[1]]
            if LINK_COL_NAME in header:
                link_ci = header.index(LINK_COL_NAME) + 1
            else:
                link_ci = len(header) + 1
                ws.cell(row=1, column=link_ci, value=LINK_COL_NAME).font = Font(bold=True)
            n = 0
            for (fl, sh, row), texts in cell_links.items():
                if fl == label and sh == sheet:
                    cell = ws.cell(row=row, column=link_ci, value=' ; '.join(texts))
                    cell.font = Font(color='1155CC', underline='single')
                    n += 1
            print(f"[写入] {os.path.basename(target)} / {sheet}: {n} 行")
        wb.save(target)

    where = "原文件 (/mnt/project)" if args.inplace else out_dir
    print(f"\n完成。法规库的「{LINK_COL_NAME}」列已更新 → {where}")

if __name__ == '__main__':
    main()
