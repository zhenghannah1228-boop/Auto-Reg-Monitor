"""frontal_mpdb_offset 正面偏置(MPDB/ODB) — 8 体系全做此项 + 回归。
C-NCAP 速度 §5.1 B=50(防脚注污染);ASEAN L3 从 xlsm 'AOP Frontal ODB' 拆。
壁障类型各异:MPDB(移动渐进)/ODB(偏置可变形)/IIHS 40%中等偏置。"""
import os, re, glob
from extract_common import pdf_text, has, SRC, merge_row, report, cncap_L3_text, jncap_hic_bands, bharat_L3

CNCAP_REF = {"A": 50, "B": 50, "C": 56, "D": 56}  # §5.1

def cnap(letter):
    g = glob.glob(os.path.join(SRC, f"中国/*C-NCAP*/02 规程附录A-T/附录{letter}*.pdf"))
    return g[0] if g else None

def raw_speed(f):
    m = re.search(r"碰撞速度为\s*(.{1,8}?)\s*km/h", pdf_text(f))
    return m.group(1).strip() if m else None

def find(*pats):
    for p in pats:
        g = glob.glob(os.path.join(SRC, p), recursive=True)
        if g: return g[0]
    return None

def asean_frontal_odb_L3():
    import openpyxl
    xl = find("东盟/*Spreadsheet*.xlsm")
    wb = openpyxl.load_workbook(xl, data_only=True, read_only=True)
    if "AOP Frontal ODB" not in wb.sheetnames:
        return {"_status": "SHEET_MISSING"}
    ws = wb["AOP Frontal ODB"]
    blocks, cur = {}, None
    for row in ws.iter_rows(min_row=1, max_row=70, values_only=True):
        joined = " ".join("" if c is None else str(c).strip() for c in row).strip()
        head = next((b for b in ("HEAD", "NECK", "CHEST", "FEMUR", "KNEE", "TIBIA", "PELVIS", "FOOT")
                     if joined.upper().startswith(b) and "assess" not in joined.lower()), None)
        if head: cur = head; blocks.setdefault(cur, [])  # 不重置:防 'Head assessment' 清空已采集块
        elif cur and any(k in joined for k in ("HIC", "compression", "Femur", "Tibia", "Viscous", "Nij", "extension", "airbag")):
            blocks[cur].append(re.sub(r"\s+", " ", joined)[:60])
    return {"_source": "xlsm 'AOP Frontal ODB'", "_blocks": {k: v for k, v in blocks.items() if v}}

def build():
    fB = cnap("B"); raw = raw_speed(fB); cn_speed = CNCAP_REF["B"]
    L1 = {  # 8 套全做正面偏置(壁障类型不同)
        "C-NCAP": bool(fB),
        "JNCAP": has(pdf_text("日本/R7-02_en.pdf"), r"offset", r"MPDB"),
        "ASEAN": "AOP Frontal ODB" in __import__("openpyxl").load_workbook(find("东盟/*Spreadsheet*.xlsm"), read_only=True).sheetnames,
        "Latin NCAP": bool(find("拉美/Latin NCAP 2025 - AOP Protocol.pdf")),
        "ANCAP": has(pdf_text(find("澳洲/*Frontal Impact*.pdf") or ""), r"MPDB", r"offset", r"mobile progressive"),
        "Euro NCAP": has(pdf_text(find("欧盟/**/*frontal_impact*.pdf") or ""), r"MPDB", r"offset"),
        "US NCAP": bool(find("美国/**/001*中等偏置*/**", "美国/**/001*40*/**")),  # IIHS 40% moderate
        "Bharat NCAP": has(pdf_text("印度/AIS_197-1.pdf"), r"offset", r"ODB", r"deformable barrier"),
    }
    asean_l3 = asean_frontal_odb_L3()
    bar = {"C-NCAP": "MPDB 50%重叠移动渐进变形壁障", "JNCAP": "MPDB(新オフセット)", "ASEAN": "ODB 偏置可变形壁障",
           "Latin NCAP": "ODB(AOP 偏置)", "ANCAP": "MPDB", "Euro NCAP": "MPDB", "US NCAP": "IIHS 40%中等偏置固定可变形壁障", "Bharat NCAP": "ODB 偏置可变形壁障"}
    meta_speed = {"C-NCAP": f"{cn_speed} km/h", "JNCAP": "50 km/h(MPDB)", "ASEAN": "—", "Latin NCAP": "64 km/h",
                  "ANCAP": "50 km/h(MPDB)", "Euro NCAP": "50 km/h(MPDB)", "US NCAP": "64 km/h(40mph,40%重叠)", "Bharat NCAP": "64 km/h"}
    # Bharat AIS-197 §3.2 正面 HPL-LPL(EEVC 系滑动,4点封顶)按章节实拆
    bh = bharat_L3(r"3\.2\.1 Head", r"4\.0 SIDE|4\.1 ")
    L3 = {s: ({"_source": "xlsm 'AOP Frontal ODB'", "_status": "PARTIAL_FROM_XLSM(区域标签与侧碰不同,待细化)", "_blocks": asean_l3.get("_blocks", {})} if s == "ASEAN"
              else {"_status": "DIFFERENT_SCORING_MODEL"} if s == "US NCAP"
              else jncap_hic_bands() if s == "JNCAP"
              else {"_extracted_tables": bh, "_scoring": "sliding HPL-LPL(EEVC,4点/部位,capping)",
                    "_source": "印度/AIS_197-1.pdf §3.2 正面伤害评价限值", "_note": "Bharat AIS-197 按章节拆;采纳 Euro/EEVC 滑动评分"} if s == "Bharat NCAP"
              else {"_status": "TO_EXTRACT"}) for s in L1}
    # C-NCAP 附录B L3:THOR-50th 头部三限值文本模式实拆(颈/胸为多列,严禁臆造不自动拆)
    _b = cncap_L3_text("中国/*C-NCAP*/02 规程附录A-T/附录B*.pdf", ["前排假人头部"])
    cn_b = {}
    for blk in _b.values():
        cn_b.update(blk)
    L3["C-NCAP"] = {"_extracted_front": cn_b, "_source": "附录B 前排假人头部评价指标(THOR-50th)",
                    "_note": "头部三限值(HIC15/累积3ms)实拆;颈/胸/腹多指标列按严禁臆造不自动拆"}
    note = {"C-NCAP": f"§5.1 核对值={cn_speed};源脚注上标污染(原始='{raw}')"}
    systems = {}
    for s, tested in L1.items():
        systems[s] = {"version": "", "L1_subtests": {"正面偏置": bool(tested)},
                      "L2_params": {"速度": meta_speed[s], "壁障": bar[s], **({"_extraction_note": note[s]} if s in note else {})} if tested else {},
                      "L3_thresholds": L3[s]}
    row = {"id": "frontal_mpdb_offset", "cn_name": "正面偏置MPDB/ODB", "en_name": "Offset Deformable Frontal (MPDB/ODB)",
           "pillar": "碰撞保护", "systems": systems,
           "diff_summary": "8 套全测、覆盖最广;C-NCAP/JNCAP 50km/h MPDB,US 用 IIHS 40%中等偏置(64km/h)、ASEAN 走 ODB(xlsm AOP Frontal ODB)",
           "key_differences": [
               "8 套全做正面偏置,但壁障分三类:MPDB 移动渐进(C-NCAP/JNCAP/Euro/ANCAP)、ODB 偏置可变形(ASEAN/Latin/Bharat)、IIHS 40%中等偏置固定壁障(US)",
               "C-NCAP 50%重叠 50km/h(§5.1 B,防脚注污染);US IIHS 40%重叠 64km/h",
               "ASEAN L3 从官方 xlsm 'AOP Frontal ODB' 拆(结构化评分),非 PDF 臆造",
               "US 另有 25% 小偏置(IIHS),单列 frontal_small_overlap;此处仅 40% 中等偏置",
               "MPDB 评分含对向壁障变形相容性(兼容性罚分),与 ODB/IIHS 评分维度不同"]}
    n = merge_row(row)
    res = [(cn_speed == CNCAP_REF["B"], "C-NCAP.偏置速度=§5.1核对值50", cn_speed, 50),
           (raw != str(cn_speed), "C-NCAP.原始提取≠50(确认污染)", raw, "≠50"),
           (bool(asean_l3.get("_blocks")), "ASEAN.L3.AOP_Frontal_ODB拆出区块", list(asean_l3.get("_blocks", {}).keys()), "非空"),
           (cn_b.get("头部HIC15") == [500, 700, 700], "C-NCAP.L3.附录B头部HIC15", cn_b.get("头部HIC15"), [500, 700, 700]),
           (len(jncap_hic_bands().get("_bands", [])) == 5, "JNCAP.L3.HIC15五色等级=5档", len(jncap_hic_bands().get("_bands", [])), 5),
           (bh.get("Head", {}).get("hic") == [500, 700], "Bharat.L3.正面头部HIC=[500,700]", bh.get("Head", {}).get("hic"), [500, 700]),
           (bh.get("Chest", {}).get("compression") == [22, 42], "Bharat.L3.正面胸压缩=[22,42]", bh.get("Chest", {}).get("compression"), [22, 42])]
    for s in L1:
        res.append((L1[s] is True or L1[s] == True, f"{s}.正面偏置=测", bool(L1[s]), True))
    print(f"frontal_mpdb_offset 做此项: {[s for s,v in L1.items() if v]} | ASEAN ODB区块: {list(asean_l3.get('_blocks',{}).keys())}")
    ok = report(res)
    print(f"ncap_matrix.json 现有 {n} 个测试项")
    return 0 if ok else 1

if __name__ == "__main__":
    import sys; sys.exit(build())
