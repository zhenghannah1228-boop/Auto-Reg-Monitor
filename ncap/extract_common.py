"""NCAP 提取公共工具:源路径、文字层读取、sample.json 回归断言。
严禁臆造:提取不到的层标 status,不编数值。"""
import os, re, json, sys
HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "sources")

# 各体系协议的【适用/生效时段】(非发布版本号);merge_row 统一写入每格 version 字段,
# 与视图顶部版本栏/报告版本表口径一致。来源依据见 coverage_report.md 顶部表。
SYSTEM_PERIOD = {
    "C-NCAP": "2027版(2027-07-01 起施行)",
    "JNCAP": "2025年度(令和7)· 年度滚动",
    "ASEAN": "2026–2030",
    "Latin NCAP": "2026–2029(实施2026.1)",
    "ANCAP": "2026–2028",
    "Euro NCAP": "2026–2028(实施2026.1)",
    "US NCAP": "按车型年度·滚动",
    "Bharat NCAP": "2023起(AIS-197/100)",
}

def _flush(pg):
    """释放 pdfplumber 单页缓存,防大 PDF(R7-03/AIS-197 等)逐页累积撑爆内存。"""
    try:
        pg.flush_cache()
    except Exception:
        pass

def pdf_text(rel, pages=None):
    import pdfplumber
    p = rel if os.path.isabs(rel) else os.path.join(SRC, rel)
    with pdfplumber.open(p) as pdf:
        ps = pdf.pages if pages is None else pdf.pages[pages[0]:pages[1]]
        parts = []
        for pg in ps:
            parts.append(pg.extract_text() or "")
            _flush(pg)          # 逐页释放:R7-03 由 538MB→79MB
        return "\n".join(parts)

def has(txt, *pats):
    """文字层是否含任一正则(大小写不敏感),用于 L1 子测试项布尔判定。"""
    return any(re.search(p, txt, re.I) for p in pats)

def load_sample():
    with open(os.path.join(HERE, "side_impact_sample.json"), encoding="utf-8") as f:
        return json.load(f)["test_item"]

def assert_eq(system, got, want, results):
    """逐字段比对(L1 布尔/L2 标量),记录到 results。"""
    for k, wv in want.items():
        gv = got.get(k)
        ok = gv == wv
        results.append((ok, f"{system}.{k}", gv, wv))

def report(results):
    bad = [r for r in results if not r[0]]
    for ok, key, gv, wv in results:
        print(("  ✓ " if ok else "  ✗ ") + key + ("" if ok else f"  got={gv!r} want={wv!r}"))
    print(f"\n{'PASS' if not bad else 'FAIL'} — {len(results)-len(bad)}/{len(results)} 字段匹配")
    return not bad


def _src_glob(gl):
    import glob
    g = glob.glob(os.path.join(SRC, gl), recursive=True)
    return g[0] if g else None

def num1(s):
    """从单元格取唯一干净数值;多 token / 公式(×≥≤△括号) / 空 → None(拒绝臆造garbled)。"""
    s = str(s or "").replace("\n", "")
    if any(x in s for x in ["×", "≥", "≤", "△", "(", ")", "（", "）"]):
        return None
    raw = re.findall(r"-?\d+(?:\.\d+)?", s)      # 带空格判定是否多 token(PDF 串列污染)
    if len(raw) != 1:
        return None
    v = raw[0]
    return float(v) if "." in v else int(v)

def cncap_L3_tables(gl):
    """C-NCAP 表模式 L3:抽『高性能限值/低性能限值』阈值表 → {组:{指标:[高,低]}}。
    附录 G/H/J 等结构表适用;garbled/公式单元格由 num1 拒绝(严禁臆造)。"""
    import pdfplumber
    f = _src_glob(gl)
    if not f:
        return {}
    out = {}
    with pdfplumber.open(f) as pdf:
        for pg in pdf.pages:
            tbs = pg.extract_tables()
            _flush(pg)
            for tb in tbs:
                hdr = [str(c or "").replace("\n", "") for c in (tb[0] or [])]
                if "高性能限值" not in " ".join(hdr):
                    continue
                hi = next((i for i, c in enumerate(hdr) if "高性能" in c), None)
                lo = next((i for i, c in enumerate(hdr) if "低性能" in c), None)
                if hi is None:
                    continue
                has_grp = hi >= 2     # 有独立分组列(col0=组,如附录J 驾驶员/二排座椅);否则 col0 即指标
                grp = "_"
                for r in tb[1:]:
                    cells = [str(c or "").replace("\n", "") for c in r]
                    if has_grp and cells and cells[0].strip() and not re.fullmatch(r"[\d.\s]+", cells[0]):
                        grp = cells[0].strip()
                    lab = ""
                    lo_i = 1 if has_grp else 0    # 有组列时,指标从 col1 起找
                    for i in range(lo_i, min(hi, len(cells))):
                        c = cells[i].strip()
                        if c and not re.fullmatch(r"[\d.\s]+", c):
                            lab = c
                    hv = num1(cells[hi]) if hi < len(cells) else None
                    lv = num1(cells[lo]) if lo is not None and lo < len(cells) else None
                    if lab and hv is not None:
                        out.setdefault(grp, {})[lab] = [hv, lv]
    return out

def cncap_L3_text(gl, captions):
    """C-NCAP 文本模式 L3:阈值以文本行渲染(附录A/B 头部表等)。给定『表X.n』标题,
    抓其后到下一『表』之间、形如『指标 高 低 极限』的单值行(num1 拒绝多列 garbled)。"""
    f = _src_glob(gl)
    if not f:
        return {}
    txt = pdf_text(f)
    lines = [l.strip() for l in txt.split("\n")]
    out = {}
    NUM = r"-?\d+(?:\.\d+)?"
    # 单值列阈值行: 指标名 (/)? 高 低 极限?  —— 三段均单一数字才采纳;指标名须含已知关键词
    KW = ("HIC", "加速度", "Fx", "Fz", "My", "压缩", "粘性", "VC", "NIC",
          "压力", "指数", "位移", "弯矩", "剪切", "张力", "合力", "耻骨", "TI", "力")
    pat = re.compile(r"^(.+?)\s+(?:/\s+)?(" + NUM + r")\s+(" + NUM + r")(?:\s+(" + NUM + r"))?\s*$")
    for cap in captions:
        try:
            start = next(i for i, l in enumerate(lines) if cap in l)
        except StopIteration:
            continue
        block = {}
        for l in lines[start + 1:start + 14]:
            if l.startswith("表") and cap not in l:
                break
            m = pat.match(l)
            if not m:
                continue
            label = m.group(1).strip()
            if not any(k in label for k in KW):
                continue
            g = m.groups()
            vals = [float(x) if (x and "." in x) else (int(x) if x else None) for x in g[1:]]
            block[label] = [v for v in vals if v is not None]
        if block:
            out[cap] = block
    return out


def cncap_L3_paired(gl):
    """C-NCAP 配对模式 L3:阈值以『高性能限值:指标 值 / 低性能限值:指标 值』两行分列
    (附录O 腿型大腿弯矩等)→ {指标:[高,低]}。仅采纳高/低同指标配对者。"""
    f = _src_glob(gl)
    if not f:
        return {}
    txt = pdf_text(f)
    NUM = r"(\d+(?:\.\d+)?)"
    hp = dict(re.findall(r"高性能限值[:：]\s*([一-龥A-Za-z/]+?)\s*" + NUM, txt))
    lp = dict(re.findall(r"低性能限值[:：]\s*([一-龥A-Za-z/]+?)\s*" + NUM, txt))
    out = {}
    for k in hp:
        if k in lp:
            hv = float(hp[k]) if "." in hp[k] else int(hp[k])
            lv = float(lp[k]) if "." in lp[k] else int(lp[k])
            out[k] = [hv, lv]
    return out


def jncap_hic_bands():
    """JNCAP 头部 HIC15 五色等级评分(2025_en.pdf 评价方法,Green→Red 五档对应 1.00→0.00 点)。
    JNCAP 独有第四类评分模型(色带);颈/胸/大腿用评价关数图(图23 等)非离散阈值,不抽。
    返回 [{band,range,points}],并带回归锚点(边界 650/1000/1350/1700、点 1.00..0.00)。"""
    txt = pdf_text("日本/2025_en.pdf")
    blk = txt[txt.find("グリーン"):txt.find("グリーン") + 1200]
    pts = re.findall(r"(\d\.\d{2})\s*点", blk)[:5]
    bnds = sorted(set(int(x) for x in
                      re.findall(r"(\d{3,4})\s*(?:≦|＜|<)\s*HIC", blk) +
                      re.findall(r"HIC\s*15?\s*(?:＜|<)\s*(\d{3,4})", blk)))
    if len(pts) != 5 or len(bnds) != 4:
        return {}
    b = bnds
    ranges = [f"<{b[0]}", f"{b[0]}–{b[1]}", f"{b[1]}–{b[2]}", f"{b[2]}–{b[3]}", f"≥{b[3]}"]
    colors = ["Green", "Yellow", "Orange", "Brown", "Red"]
    bands = [{"band": c, "range": r, "points": float(p)} for c, r, p in zip(colors, ranges, pts)]
    return {"_scoring": "5色等级(Green→Red,色带评分)", "_metric": "头部 HIC15",
            "_source": "日本/2025_en.pdf 自動車安全性能評価方法(別添 判定基準)",
            "_bands": bands,
            "_note": "JNCAP 头部 HIC15 五色等级实拆;颈/胸/大腿部用评价关数(评价函数图,如图23)连续评分,非离散阈值,按严禁臆造不抽"}


def bharat_L3(start_re, end_re):
    """Bharat AIS-197 按章节拆 L3:Euro/EEVC 系滑动 HPL-LPL(每部位满分4点,capping 封顶)。
    解析『Higher/Lower performance limit』块 → {部位:{指标:[HPL,LPL]}};去 HIC15 下标/3msec
    噪声,仅采纳 HPL+LPL 均解析出的指标(严禁臆造)。section 由起止正则界定。"""
    txt = pdf_text("印度/AIS_197-1.pdf")
    lines = [l.strip() for l in txt.split("\n")]
    try:
        si = next(i for i, l in enumerate(lines) if re.search(start_re, l))
    except StopIteration:
        return {}
    ei = next((i for i in range(si + 1, len(lines)) if re.search(end_re, lines[i])), len(lines))
    out, reg, mode = {}, None, None
    for l0 in lines[si:ei]:
        m = re.match(r"\d\.\d\.\d\.?\s+(Head|Neck|Chest|Abdomen|Pelvis|Knee|Femur)", l0)
        if m:
            reg = m.group(1); out.setdefault(reg, {}); continue
        if re.search(r"[Hh]igher performance", l0):
            mode = "HPL"; continue
        if re.search(r"[Ll]ower performance", l0):
            mode = "LPL"; continue
        if not (reg and mode):
            continue
        l = re.sub(r"HIC\s*15", "HIC", l0)
        l = re.sub(r"3\s*m?sec", "", l)
        mm = re.match(r"([A-Za-z][A-Za-z .]+?)\s*(\d+(?:\.\d+)?)\s*(kN|mm|g|m/sec)?", l)
        if mm:
            ind = re.sub(r"\s+", " ", mm.group(1).strip()).lower()
            val = float(mm.group(2)) if "." in mm.group(2) else int(mm.group(2))
            out[reg].setdefault(ind, {})[mode] = val
    res = {}
    for reg, inds in out.items():
        d = {k: [v["HPL"], v["LPL"]] for k, v in inds.items() if "HPL" in v and "LPL" in v}
        if d:
            res[reg] = d
    return res


def asean_fitment(sheet):
    """ASEAN 主动安全『装备率评分』(Fitment Rating)实拆:Option A/B/C…→α 分值 + TFS 总分。
    业主 2026-06 确认:ASEAN 主动安全按是否装配打分(非性能阈值)——独立第五类评分模型,
    『不测性能只看装配』本身即高价值跨体系差集。表 col0 空,需全行扫描。"""
    import openpyxl
    xl = _src_glob("东盟/*Spreadsheet*.xlsm")
    if not xl:
        return {}
    wb = openpyxl.load_workbook(xl, data_only=True, read_only=True)
    sn = next((s for s in wb.sheetnames if s.strip() == sheet.strip()), None)
    if not sn:
        return {}
    ws = wb[sn]
    title, options, tfs = None, [], None
    for r in ws.iter_rows(min_row=1, max_row=26, values_only=True):
        cells = [("" if c is None else str(c).strip()) for c in r]
        for c in cells:
            if "Fitment Rating System" in c:
                title = c
        nums = [c for c in cells if re.fullmatch(r"\d+(?:\.\d+)?", c)]
        opt = next((c for c in cells if re.fullmatch(r"Option [A-G]", c)), None)
        di = next((i for i, c in enumerate(cells) if len(c) > 12 and "equipped" in c.lower()), None)
        # 仅采纳『定义表』行(含 equipped 描述);α = detail 之后第一个数字(避开相邻测试表串入的污染)
        if opt and di is not None:
            alpha = next((float(c) for c in cells[di + 1:] if re.fullmatch(r"\d+(?:\.\d+)?", c)), None)
            if alpha is not None:
                options.append({"option": opt, "alpha": alpha,
                                "detail": re.sub(r"\s+", " ", cells[di])[:70]})
        if "TFS" in cells and nums:
            tfs = float(nums[-1])
    if not options:
        return {}
    return {"_scoring": "装备率(fitment:标配α/选配/无 × 国别系数)", "_fitment": options,
            "_TFS": tfs, "_title": title, "_source": f"东盟 NCAP xlsm sheet '{sn.strip()}'",
            "_note": "ASEAN 主动安全按整车是否装配该功能评分(Option A标配→α满分/B选配/C无→0),"
                     "非性能阈值——『不测性能、只看装配』是与中/欧/日/印的根本差集"}


def euro_hpl_lpl(file_glob, pages, criteria):
    """Euro NCAP AOP 滑动 HPL-LPL 阈值实拆(色带 Green1.25/Yellow1.0/Orange0.75/Brown0.5/Red0)。
    多假人列按 [HIII5th, HIII50th, THOR50th, HIII95th] 排列,取 **HIII 50th=第2个 num-num 对**
    (标准成人);续行用 3 行窗口合并。criteria={显示名:行内正则}。Latin/ANCAP 采纳同表。"""
    import pdfplumber
    f = _src_glob(file_glob)
    if not f:
        return {}
    with pdfplumber.open(f) as pdf:
        lines = []
        for pg in pdf.pages:
            if pages is None or pg.page_number in pages:
                lines += (pg.extract_text() or "").split("\n")
            _flush(pg)
    PAIR = re.compile(r"(\d+(?:\.\d+)?)\s*-\s*(\d+(?:\.\d+)?)")
    num = lambda x: float(x) if "." in x else int(x)
    out = {}
    for name, pat in criteria.items():
        for i, ln in enumerate(lines):
            if re.search(pat, ln):
                pairs = PAIR.findall(" ".join(lines[i:i + 3]))
                if len(pairs) >= 2:
                    out[name] = [num(pairs[1][0]), num(pairs[1][1])]
                break
    return out

EURO_BANDS_NOTE = "色带:Green 1.25/Yellow 1.0/Orange 0.75/Brown 0.5/Red 0 分;HPL=满分边界、LPL=0分边界,线性内插,capping 封顶"

def euro_pedestrian_head_bands():
    """Euro 行人头型 HIC15 五色等级(Pedestrian v8.5 §4 Headform Data:Green<650…Red≥1700)。
    Latin/ANCAP 采纳同表。腿型(legform)用 UN R127 弯矩/韧带,评分限值在图/网格,不抽。"""
    f = _src_glob("拉美/Euro NCAP - Pedestrian*8.5*.pdf")
    if not f:
        return {}
    txt = pdf_text(f).replace("\n", " ")
    # 定位含完整 5 色梯(Green…Red)的段:扫描每个『= Green』窗口,取同时含 Brown+Red 者
    seg = None
    for m in re.finditer(r"=\s*Green", txt):
        w = txt[m.start() - 40:m.start() + 360]
        if "Brown" in w and "Red" in w:
            seg = w; break
    if not seg:
        return {}
    bnds = sorted(set(int(x) for pair in re.findall(r"(\d{3,4})\s*≤\s*HIC|HIC\s*<\s*(\d{3,4})", seg) for x in pair if x))
    colors = re.findall(r"=\s*(Green|Yellow|Orange|Brown|Red)", seg)[:5]
    if len(bnds) != 4 or len(colors) != 5:
        return {}
    b = bnds
    ranges = [f"<{b[0]}", f"{b[0]}–{b[1]}", f"{b[1]}–{b[2]}", f"{b[2]}–{b[3]}", f"≥{b[3]}"]
    return {"_scoring": "5色等级(头型 HIC15 网格,色带)", "_metric": "头型 HIC15",
            "_source": "拉美/Euro NCAP Pedestrian v8.5 §4 Headform Data",
            "_bands": [{"band": c, "range": r} for c, r in zip(colors, ranges)],
            "_note": "Euro/Latin/ANCAP 行人头型 HIC15 五色网格;腿型(UN R127 胫骨弯矩/韧带)评分限值在图/网格,不抽"}


def bharat_ais100_pedestrian():
    """Bharat 行人保护遵循 AIS-100(UN GTR9 系)合规限值 pass-fail(非 NCAP 滑动/色带)。
    §5 性能要求:头型 HIC1000/1700 区、下腿型 MCL≤22/ACL·PCL≤13/胫骨弯矩≤340、
    上腿型 合力≤7.5kN/弯矩≤510Nm。单上限值(≤),实拆 + 锚点回归。"""
    f = _src_glob("印度/AIS_100.pdf")
    if not f:
        return {}
    t = pdf_text(f).replace("\n", " ")
    def g(pat):
        m = re.search(pat, t, re.I)
        return (float(m.group(1)) if "." in m.group(1) else int(m.group(1))) if m else None
    mcl = g(r"medial collateral\s*ligament elongation at the knee shall not exceed\s*(\d+)\s*mm")
    acl = g(r"posterior cruciate\s*ligament elongation shall not exceed\s*(\d+)\s*mm")
    tib = g(r"bending moments at the tibia shall not exceed\s*(\d+)\s*Nm")
    upf = g(r"sum\s*of\s*the impact forces.*?not exceed\s*([\d.]+)\s*kN")
    upm = g(r"bending moment on the test impactor shall not exceed\s*(\d+)\s*Nm")
    has1000 = bool(re.search(r"HIC recorded shall not exceed\s*1,?000", t))
    has1700 = bool(re.search(r"not exceed\s*1,?700", t))
    tables = {
        "头型 Headform": {"HIC1000区(≤)": [1000] if has1000 else [], "HIC1700区其余(≤)": [1700] if has1700 else []},
        "下腿型 LowerLeg": {"MCL韧带(mm≤)": [mcl], "ACL·PCL韧带(mm≤)": [acl], "胫骨弯矩(Nm≤)": [tib]},
        "上腿型 UpperLeg": {"合力(kN≤)": [upf], "弯矩(Nm≤)": [upm]},
    }
    return {"_extracted_tables": tables, "_scoring": "pass-fail 合规限值(AIS-100 / UN GTR9,单上限≤)",
            "_source": "印度/AIS_100.pdf §5 性能要求(Rev.1)",
            "_note": "Bharat 行人保护按 AIS-100 法规合规限值(单上限 pass-fail,非 NCAP 滑动/色带评分);"
                     "头型分 HIC1000 区(≥半区)与 HIC1700 区(其余)"}


def merge_row(row):
    """把一个测试项行并入 ncap_matrix.json(按 id 去重替换),数组形式。
    统一把每格 version 写成【适用/生效时段】(SYSTEM_PERIOD),与视图版本栏口径一致;
    原发布版本号(子协议)保留在 source_files / L2._note 中。"""
    import json
    for s, obj in (row.get("systems") or {}).items():
        if s in SYSTEM_PERIOD and isinstance(obj, dict):
            obj["version"] = SYSTEM_PERIOD[s]
    p = os.path.join(HERE, "ncap_matrix.json")
    arr = []
    if os.path.exists(p):
        arr = json.load(open(p, encoding="utf-8"))
    arr = [x for x in arr if x.get("id") != row.get("id")] + [row]
    # 固定顺序:碰撞保护族在前,稳定输出
    order = ["frontal_rigid_full", "frontal_mpdb_offset", "frontal_small_overlap",
             "side_impact", "side_pole", "whiplash_rear", "post_crash_safety",
             "ev_hazard", "restraint_system", "vru_passive",
             "adas_aeb", "vru_active", "lane_support", "blind_spot",
             "adaptive_highbeam", "occupant_monitoring"]
    arr.sort(key=lambda x: order.index(x["id"]) if x["id"] in order else 99)
    json.dump(arr, open(p, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    return len(arr)
